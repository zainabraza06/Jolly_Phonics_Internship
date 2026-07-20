"""
unified_pipeline.py — single source of truth for the phoneme-gesture model.

Does everything in one run:
  1. Loads the trained FusionClassifier from checkpoint (one model definition,
     used everywhere below — no more copy-pasted classes that can drift).
  2. Builds (or loads cached) reference stats for BOTH elder and child groups
     -> reference_stats.pkl
  3. Runs expert evaluation over expert_test_videos/:
       - classifies each clip using the REAL temporal sequence (fixes the bug
         where audio/video features were mean-pooled then .expand()'d into a
         fake flat sequence, which destroyed the BiLSTM's temporal signal)
       - scores each clip against the matching reference group using the
         pooled distance-based scorer
       - writes expert_evaluation.xlsx with blank expert columns
  4. Exports an app-integration bundle:
       model_export/
         model.pth
         model_config.json
         label_map.json
         reference_stats.pkl      (child + elder, for on-device scoring)
         inference.py             (predict() AND score(), self-contained)

Run once in Colab. Everything downstream (web app backend) only needs
model_export/.
"""

import subprocess, sys
subprocess.run([sys.executable, "-m", "pip", "install", "-q",
                "torch", "torchaudio", "transformers", "mediapipe",
                "noisereduce", "soundfile", "opencv-python-headless",
                "openpyxl", "numpy"])

from google.colab import drive
drive.mount('/content/drive')

import os, re, json, pickle, shutil, urllib.request
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import torchaudio
import cv2
import soundfile as sf
import noisereduce as nr
from transformers import WhisperProcessor, WhisperModel
from mediapipe.tasks import python as mp_python
from mediapipe.tasks.python import vision as mp_vision
import mediapipe as mp
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIG / PATHS — edit these
# =============================================================================
MODEL_SAVE_ROOT = "/content/drive/MyDrive/phoneme_gesture_ablations/ablation_checkpoints"
CHECKPOINT_NAME = "checkpoint_cmp_ablation20_fp32_cross_attn_bce_bilstm_attn_epoch30.pth"
CHECKPOINT_PATH = os.path.join(MODEL_SAVE_ROOT, CHECKPOINT_NAME)

# unimodal checkpoints — used to add audio-only / video-only columns to the
# Excel sheet alongside the fusion model's predictions. Set to None to skip
# either (its Excel columns will just be left blank).
AUDIO_ONLY_CHECKPOINT_NAME = "ablation9_fp32_audio_only_bce_best.pth"
AUDIO_ONLY_CHECKPOINT_PATH = os.path.join(MODEL_SAVE_ROOT, AUDIO_ONLY_CHECKPOINT_NAME)

VIDEO_ONLY_CHECKPOINT_NAME = "checkpoint_cmp_ablation12_fp32_video_only_bce_bilstm_attn_epoch30.pth"
VIDEO_ONLY_CHECKPOINT_PATH = os.path.join(MODEL_SAVE_ROOT, VIDEO_ONLY_CHECKPOINT_NAME)

TEST_DIR       = "/content/drive/MyDrive/experts_test_data"
REF_ELDER_DIR  = "/content/drive/MyDrive/references_elder"
REF_CHILD_DIR  = "/content/drive/MyDrive/references_child"

APP_ROOT       = "/content/drive/MyDrive/app_integration"
EXPORT_DIR     = os.path.join(APP_ROOT, "model_export")
HF_CACHE       = os.path.join(EXPORT_DIR, "hf_cache")
REF_STATS_PKL  = os.path.join(EXPORT_DIR, "reference_stats.pkl")
EXCEL_OUT      = "/content/drive/MyDrive/expert_evaluation.xlsx"

TMP_DIR = "/content/tmp_pipeline"
MP_DIR  = "/content/mp_models"
os.makedirs(EXPORT_DIR, exist_ok=True)
os.makedirs(HF_CACHE, exist_ok=True)
os.makedirs(TMP_DIR, exist_ok=True)
os.makedirs(MP_DIR, exist_ok=True)

REBUILD_REF_STATS = True    # IMPORTANT: force a rebuild this run — the cached
                             # reference_stats.pkl was built before the 'ck'
                             # filename fix and before self-calibrated scoring
                             # thresholds existed, so it's stale. Once you've
                             # run this and confirmed the new pkl looks right,
                             # you can flip this back to False for future runs.

# ---- prefix -> full label (must match training) ----
PREFIX_TO_LABEL = {
    "s":   "s first letter of gp 01",
    "ai":  "ai first letter of gp 04",
    "c-k": "c-k first letter of group 02",
    "ck":  "c-k first letter of group 02",  # filenames use 'ck' with no hyphen
    "g":   "g first letter of gp 03",
    "qu":  "qu first letter of gp 07",
    "y":   "y first letter of gp 06",
    "z":   "z first letter of gp 05",
}
LABELS_SORTED = sorted(set(PREFIX_TO_LABEL.values()))
label_to_id   = {l: i for i, l in enumerate(LABELS_SORTED)}
id_to_label   = {i: l for l, i in label_to_id.items()}

WHISPER_LAYERS   = [1, 3, 5, 7, 9]
VIDEO_FRAMES     = 64
MAX_AUDIO_FRAMES = 200
VIDEO_FEAT_DIM   = 234
MOUTH_IDS = [61,146,91,181,84,17,314,405,321,375,291,308,
             324,318,402,317,14,87,178,88,95,185,40,39,37,
             0,267,269,270,409]

# scoring thresholds — DEFAULT FALLBACK ONLY. Used when a group has fewer
# than 2 reference classes for a modality, so per-class self-calibration
# (see compute_calibrated_thresholds()) isn't possible. In normal operation
# thresholds are derived automatically per class from the actual distances
# between your reference embeddings, so they self-scale to whatever the
# real feature space looks like instead of assuming a fixed distance scale.
DEFAULT_AUDIO_GOOD, DEFAULT_AUDIO_MAX, DEFAULT_AUDIO_BAD = 1.5, 4.0, 8.0
DEFAULT_LM_GOOD,    DEFAULT_LM_MAX,    DEFAULT_LM_BAD    = 0.8, 2.5, 5.0

MODEL_CONFIG = {
    "num_classes": len(LABELS_SORTED),
    "fusion_type": "cross_attn",
    "video_arch": "bilstm_attn",
    "whisper_model": "openai/whisper-small",
    "whisper_layers": WHISPER_LAYERS,
    "audio_bilstm_hidden": 512,
    "audio_bilstm_layers": 2,
    "video_bilstm_hidden": 128,
    "video_bilstm_layers": 1,
    "cross_attn_proj_dim": 256,
    "cross_attn_heads": 4,
    "classifier_hidden": 256,
    "classifier_dropout": 0.3,
    "video_feat_dim": VIDEO_FEAT_DIM,
    "video_frames": VIDEO_FRAMES,
    "audio_max_frames": MAX_AUDIO_FRAMES,
}

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


# =============================================================================
# MODEL DEFINITION — single copy, used by eval AND export
# =============================================================================
class AudioBiLSTMEncoder(nn.Module):
    def __init__(self, input_dim=768, hidden_size=512, num_layers=2, dropout=0.5):
        super().__init__()
        self.bilstm = nn.LSTM(input_dim, hidden_size, num_layers,
                               bidirectional=True, batch_first=True,
                               dropout=dropout if num_layers > 1 else 0)
        self.ln = nn.LayerNorm(hidden_size * 2)
        self.attention = nn.Sequential(
            nn.Linear(hidden_size * 2, hidden_size), nn.Tanh(),
            nn.Linear(hidden_size, 1, bias=False))
        self.out_dim = hidden_size * 2

    def forward(self, x, lengths):
        packed = nn.utils.rnn.pack_padded_sequence(
            x, lengths.cpu(), batch_first=True, enforce_sorted=False)
        out, _ = self.bilstm(packed)
        out, _ = nn.utils.rnn.pad_packed_sequence(out, batch_first=True)
        seq = self.ln(out)
        w = F.softmax(self.attention(seq).squeeze(-1), dim=1)
        pooled = torch.bmm(w.unsqueeze(1), seq).squeeze(1)
        return seq, pooled


class VideoBiLSTMEncoder(nn.Module):
    def __init__(self, input_dim=VIDEO_FEAT_DIM, hidden_size=128, num_layers=1, dropout=0.3):
        super().__init__()
        self.bilstm = nn.LSTM(input_dim, hidden_size, num_layers,
                               bidirectional=True, batch_first=True,
                               dropout=dropout if num_layers > 1 else 0)
        self.ln = nn.LayerNorm(hidden_size * 2)
        self.attention = nn.Sequential(
            nn.Linear(hidden_size * 2, hidden_size), nn.Tanh(),
            nn.Linear(hidden_size, 1, bias=False))
        self.out_dim = hidden_size * 2

    def forward(self, x, lengths=None):
        out, _ = self.bilstm(x)
        seq = self.ln(out)
        w = F.softmax(self.attention(seq).squeeze(-1), dim=1)
        pooled = torch.bmm(w.unsqueeze(1), seq).squeeze(1)
        return seq, pooled


class CrossAttentionFusion(nn.Module):
    def __init__(self, audio_dim, video_dim, proj_dim=256, num_heads=4, dropout=0.2):
        super().__init__()
        self.a_proj = nn.Linear(audio_dim, proj_dim)
        self.v_proj = nn.Linear(video_dim, proj_dim)
        self.a2v_attn = nn.MultiheadAttention(proj_dim, num_heads, dropout=dropout, batch_first=True)
        self.v2a_attn = nn.MultiheadAttention(proj_dim, num_heads, dropout=dropout, batch_first=True)
        self.ln_a = nn.LayerNorm(proj_dim)
        self.ln_v = nn.LayerNorm(proj_dim)
        self.out_dim = proj_dim * 2

    @staticmethod
    def _masked_mean(x, mask):
        valid = (~mask).unsqueeze(-1).float()
        return (x * valid).sum(1) / valid.sum(1).clamp(min=1)

    def forward(self, audio_seq, audio_mask, video_seq, video_mask):
        a = self.a_proj(audio_seq)
        v = self.v_proj(video_seq)
        v_ctx, _ = self.v2a_attn(query=v, key=a, value=a, key_padding_mask=audio_mask)
        a_ctx, _ = self.a2v_attn(query=a, key=v, value=v, key_padding_mask=video_mask)
        v_ctx = self.ln_v(v_ctx + v)
        a_ctx = self.ln_a(a_ctx + a)
        return torch.cat([self._masked_mean(a_ctx, audio_mask),
                           self._masked_mean(v_ctx, video_mask)], dim=-1)


class FusionClassifier(nn.Module):
    def __init__(self, num_classes=7, proj_dim=256, dropout=0.3):
        super().__init__()
        self.audio_enc = AudioBiLSTMEncoder(hidden_size=512, num_layers=2, dropout=0.5)
        self.video_enc = VideoBiLSTMEncoder(hidden_size=128, num_layers=1)
        self.cross_attn = CrossAttentionFusion(
            self.audio_enc.out_dim, self.video_enc.out_dim, proj_dim=proj_dim, dropout=dropout)
        fused_dim = self.cross_attn.out_dim
        self.classifier = nn.Sequential(
            nn.Linear(fused_dim, 256), nn.ReLU(), nn.Dropout(dropout),
            nn.Linear(256, num_classes))

    def forward(self, video_x, video_len, audio_x, audio_len):
        v_seq, _ = self.video_enc(video_x)
        a_seq, _ = self.audio_enc(audio_x, audio_len)
        Tv = v_seq.shape[1]
        Ta = a_seq.shape[1]
        v_mask = torch.arange(Tv, device=video_x.device).unsqueeze(0) >= video_len.to(video_x.device).unsqueeze(1)
        a_mask = torch.arange(Ta, device=audio_x.device).unsqueeze(0) >= audio_len.to(audio_x.device).unsqueeze(1)
        fused = self.cross_attn(a_seq, a_mask, v_seq, v_mask)
        return self.classifier(fused)


class AudioOnlyClassifier(nn.Module):
    """Matches ablation9 (audio_only, fp32, BCE) checkpoint exactly.

    NOTE: the checkpoint saved during training also contains a `video_enc`
    submodule (VideoBiLSTMEncoder) — it was instantiated as part of the
    training-time model class but never actually used in forward().
    Confirmed by checkpoint inspection: classifier.0.weight has shape
    (256, 1024), i.e. the classifier only ever consumed audio_enc's pooled
    1024-dim output (512 hidden * 2 directions) — never a concatenation
    with video's 256-dim output, which would be (256, 1280).

    video_enc is kept here ONLY so state_dict keys line up on load; it is
    never called in forward() and has zero effect on predictions.
    """
    def __init__(self, num_classes=7, dropout=0.3):
        super().__init__()
        self.audio_enc = AudioBiLSTMEncoder(hidden_size=512, num_layers=2, dropout=0.5)
        self.video_enc = VideoBiLSTMEncoder(hidden_size=128, num_layers=1)  # unused, load-only
        self.classifier = nn.Sequential(
            nn.Linear(self.audio_enc.out_dim, 256), nn.ReLU(), nn.Dropout(dropout),
            nn.Linear(256, num_classes))

    def forward(self, audio_x, audio_len):
        _, pooled = self.audio_enc(audio_x, audio_len)
        return self.classifier(pooled)


class VideoOnlyClassifier(nn.Module):
    """Matches ablation12 (video_only, bilstm_attn): VideoBiLSTMEncoder ->
    plain Linear(num_classes), no hidden layer (matches Main2_working.py)."""
    def __init__(self, num_classes=7):
        super().__init__()
        self.video_enc = VideoBiLSTMEncoder(hidden_size=128, num_layers=1)
        self.classifier = nn.Linear(self.video_enc.out_dim, num_classes)

    def forward(self, video_x, video_len=None):
        _, pooled = self.video_enc(video_x, video_len)
        return self.classifier(pooled)


print(f"Loading model checkpoint on {device}...")
model = FusionClassifier(num_classes=len(LABELS_SORTED)).to(device)
model.load_state_dict(torch.load(CHECKPOINT_PATH, map_location=device))
model.eval()
print("✓ Model loaded")

audio_only_model = None
if AUDIO_ONLY_CHECKPOINT_PATH and os.path.exists(AUDIO_ONLY_CHECKPOINT_PATH):
    print(f"Loading audio-only checkpoint on {device}...")
    audio_only_model = AudioOnlyClassifier(num_classes=len(LABELS_SORTED)).to(device)
    audio_only_model.load_state_dict(torch.load(AUDIO_ONLY_CHECKPOINT_PATH, map_location=device))
    audio_only_model.eval()
    print("✓ Audio-only model loaded")
else:
    print(f"  WARNING: audio-only checkpoint not found at {AUDIO_ONLY_CHECKPOINT_PATH} — "
          f"audio-only columns will be blank")

video_only_model = None
if VIDEO_ONLY_CHECKPOINT_PATH and os.path.exists(VIDEO_ONLY_CHECKPOINT_PATH):
    print(f"Loading video-only checkpoint on {device}...")
    video_only_model = VideoOnlyClassifier(num_classes=len(LABELS_SORTED)).to(device)
    video_only_model.load_state_dict(torch.load(VIDEO_ONLY_CHECKPOINT_PATH, map_location=device))
    video_only_model.eval()
    print("✓ Video-only model loaded")
else:
    print(f"  WARNING: video-only checkpoint not found at {VIDEO_ONLY_CHECKPOINT_PATH} — "
          f"video-only columns will be blank")


# =============================================================================
# WHISPER (audio feature extractor)
# =============================================================================
print("Loading Whisper-small...")
whisper_processor = WhisperProcessor.from_pretrained("openai/whisper-small", cache_dir=HF_CACHE)
whisper_model = WhisperModel.from_pretrained(
    "openai/whisper-small", output_hidden_states=True, cache_dir=HF_CACHE)
whisper_model.eval().to(device)
_layer_weights = torch.ones(len(WHISPER_LAYERS), device=device)
_audio_ln = nn.LayerNorm(768).to(device)


def _mp4_to_clean_wav(mp4_path, out_wav):
    subprocess.run(["ffmpeg", "-y", "-i", mp4_path, "-vn", "-acodec", "pcm_s16le",
                     "-ar", "16000", "-ac", "1", out_wav],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    audio, sr = sf.read(out_wav)
    audio = nr.reduce_noise(y=audio, sr=sr, stationary=True)
    sf.write(out_wav, audio, sr)


def extract_audio_sequence(wav_path):
    """Returns (T, 768) full temporal sequence — what the model was trained on."""
    wav, sr = torchaudio.load(wav_path)
    if wav.shape[0] > 1:
        wav = wav.mean(0, keepdim=True)
    if sr != 16000:
        wav = torchaudio.functional.resample(wav, sr, 16000)
    inputs = whisper_processor(wav.squeeze().numpy(), sampling_rate=16000, return_tensors="pt")
    feats = inputs.input_features.to(device)
    with torch.no_grad():
        out = whisper_model.encoder(input_features=feats)
        hidden = out.hidden_states
    selected = [hidden[i].squeeze(0) for i in WHISPER_LAYERS]
    w = F.softmax(_layer_weights, dim=0)
    combined = torch.sum(w[:, None, None] * torch.stack(selected), dim=0)
    seq = _audio_ln(combined).detach()  # (T, 768) on device
    if seq.shape[0] > MAX_AUDIO_FRAMES:
        seq = seq[:MAX_AUDIO_FRAMES]
    return seq.cpu()


# =============================================================================
# MEDIAPIPE (video/landmark feature extractor)
# =============================================================================
def _dl(p, u):
    if not os.path.exists(p):
        print(f"  Downloading {os.path.basename(p)} ...")
        urllib.request.urlretrieve(u, p)

HAND_M = os.path.join(MP_DIR, "hand_landmarker.task")
POSE_M = os.path.join(MP_DIR, "pose_landmarker_lite.task")
FACE_M = os.path.join(MP_DIR, "face_landmarker.task")
_dl(HAND_M, "https://storage.googleapis.com/mediapipe-models/hand_landmarker/hand_landmarker/float16/1/hand_landmarker.task")
_dl(POSE_M, "https://storage.googleapis.com/mediapipe-models/pose_landmarker/pose_landmarker_lite/float16/1/pose_landmarker_lite.task")
_dl(FACE_M, "https://storage.googleapis.com/mediapipe-models/face_landmarker/face_landmarker/float16/1/face_landmarker.task")


def _make_landmarkers():
    h = mp_vision.HandLandmarker.create_from_options(
        mp_vision.HandLandmarkerOptions(
            base_options=mp_python.BaseOptions(model_asset_path=HAND_M),
            running_mode=mp_vision.RunningMode.VIDEO, num_hands=2))
    p = mp_vision.PoseLandmarker.create_from_options(
        mp_vision.PoseLandmarkerOptions(
            base_options=mp_python.BaseOptions(model_asset_path=POSE_M),
            running_mode=mp_vision.RunningMode.VIDEO))
    f = mp_vision.FaceLandmarker.create_from_options(
        mp_vision.FaceLandmarkerOptions(
            base_options=mp_python.BaseOptions(model_asset_path=FACE_M),
            running_mode=mp_vision.RunningMode.VIDEO))
    return h, p, f


def _normalize_landmarks(arr):
    """arr: (T, 234) raw -> body-relative normalized, still (T, 234)."""
    T, D = arr.shape
    pose = arr[:, 126:144].reshape(T, 6, 3)
    l_sh, r_sh = pose[:, 0], pose[:, 1]
    mid = (l_sh + r_sh) / 2.0
    width = np.linalg.norm((l_sh - r_sh)[:, :2], axis=1).clip(1e-4)
    out = arr.reshape(T, D // 3, 3)
    out = (out - mid[:, None, :]) / width[:, None, None]
    return out.reshape(T, D).astype(np.float32)


def extract_video_sequence(video_path):
    """Returns (VIDEO_FRAMES, 234) full temporal sequence — resampled to a
    fixed number of frames, normalized. This is what the model expects."""
    hl, pl, fl = _make_landmarkers()
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
    rows, idx = [], 0
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        mp_img = mp.Image(image_format=mp.ImageFormat.SRGB, data=rgb)
        ts = int(idx / fps * 1000)
        hr = hl.detect_for_video(mp_img, ts)
        pr = pl.detect_for_video(mp_img, ts)
        fr = fl.detect_for_video(mp_img, ts)
        hv = np.zeros((2, 21, 3), np.float32)
        for i, h_ in enumerate(hr.hand_landmarks[:2]):
            for j, lm in enumerate(h_):
                hv[i, j] = [lm.x, lm.y, lm.z]
        pv = np.zeros((6, 3), np.float32)
        if pr.pose_landmarks:
            for i, pid in enumerate([11, 12, 13, 14, 15, 16]):
                lm = pr.pose_landmarks[0][pid]; pv[i] = [lm.x, lm.y, lm.z]
        mv = np.zeros((len(MOUTH_IDS), 3), np.float32)
        if fr.face_landmarks:
            for i, fid in enumerate(MOUTH_IDS):
                lm = fr.face_landmarks[0][fid]; mv[i] = [lm.x, lm.y, lm.z]
        rows.append(np.concatenate([hv.flatten(), pv.flatten(), mv.flatten()]))
        idx += 1
    cap.release()
    for x in [hl, pl, fl]:
        x.close()

    arr = np.stack(rows) if rows else np.zeros((1, VIDEO_FEAT_DIM), np.float32)
    T, D = arr.shape
    if T != VIDEO_FRAMES:
        src = np.linspace(0, T - 1, T)
        tgt = np.linspace(0, T - 1, VIDEO_FRAMES)
        out = np.zeros((VIDEO_FRAMES, D), np.float32)
        for d in range(D):
            out[:, d] = np.interp(tgt, src, arr[:, d])
        arr = out
    arr = _normalize_landmarks(arr)
    return arr  # (VIDEO_FRAMES, 234)


# =============================================================================
# CLASSIFICATION — uses REAL sequences (this is the fix vs. the old script)
# =============================================================================
def classify(audio_seq_np_or_t, video_seq_np):
    """
    audio_seq: (T, 768) real temporal sequence (torch.Tensor or np.ndarray)
    video_seq: (VIDEO_FRAMES, 234) real temporal sequence (np.ndarray)
    """
    if isinstance(audio_seq_np_or_t, np.ndarray):
        audio_seq_np_or_t = torch.from_numpy(audio_seq_np_or_t)
    audio_x = audio_seq_np_or_t.float().unsqueeze(0).to(device)          # (1, T, 768)
    video_x = torch.from_numpy(video_seq_np).float().unsqueeze(0).to(device)  # (1, 64, 234)
    audio_len = torch.tensor([audio_x.shape[1]])
    video_len = torch.tensor([video_x.shape[1]])

    with torch.no_grad():
        logits = model(video_x, video_len, audio_x, audio_len)
        probs = torch.sigmoid(logits).squeeze(0).cpu().numpy()

    pred_id = int(probs.argmax())
    pred_label = id_to_label[pred_id]
    confidence = round(float(probs[pred_id]) * 100, 2)
    class_probs = {id_to_label[i]: round(float(p) * 100, 2) for i, p in enumerate(probs)}
    return pred_label, confidence, class_probs


def classify_audio_only(audio_seq_np_or_t):
    """audio_seq: (T, 768) real temporal sequence. Returns (None, None, {})
    if the audio-only model isn't loaded."""
    if audio_only_model is None:
        return None, None, {}
    if isinstance(audio_seq_np_or_t, np.ndarray):
        audio_seq_np_or_t = torch.from_numpy(audio_seq_np_or_t)
    audio_x = audio_seq_np_or_t.float().unsqueeze(0).to(device)  # (1, T, 768)
    audio_len = torch.tensor([audio_x.shape[1]])
    with torch.no_grad():
        logits = audio_only_model(audio_x, audio_len)
        probs = torch.sigmoid(logits).squeeze(0).cpu().numpy()
    pred_id = int(probs.argmax())
    pred_label = id_to_label[pred_id]
    confidence = round(float(probs[pred_id]) * 100, 2)
    class_probs = {id_to_label[i]: round(float(p) * 100, 2) for i, p in enumerate(probs)}
    return pred_label, confidence, class_probs


def classify_video_only(video_seq_np):
    """video_seq: (VIDEO_FRAMES, 234) real temporal sequence. Returns
    (None, None, {}) if the video-only model isn't loaded."""
    if video_only_model is None:
        return None, None, {}
    video_x = torch.from_numpy(video_seq_np).float().unsqueeze(0).to(device)  # (1, 64, 234)
    video_len = torch.tensor([video_x.shape[1]])
    with torch.no_grad():
        logits = video_only_model(video_x, video_len)
        probs = torch.sigmoid(logits).squeeze(0).cpu().numpy()
    pred_id = int(probs.argmax())
    pred_label = id_to_label[pred_id]
    confidence = round(float(probs[pred_id]) * 100, 2)
    class_probs = {id_to_label[i]: round(float(p) * 100, 2) for i, p in enumerate(probs)}
    return pred_label, confidence, class_probs


# =============================================================================
# SCORING — pooled (mean) vectors compared to reference, distance -> 0-100,
# blended with the model's own confidence in the true label (since in
# production there's only ever a single attempt to score — no distribution
# of other attempts to calibrate against, so the model's own probability
# output is the only other independent signal available at inference time).
# =============================================================================
CONFIDENCE_BLEND_WEIGHT = 0.4  # 0 = pure distance score, 1 = pure model confidence

def _score_one(dist, good, max_, bad):
    if dist <= good:
        v = 80 + 20 * (1 - dist / good)
    elif dist <= max_:
        v = 50 + 30 * (1 - (dist - good) / (max_ - good))
    elif dist <= bad:
        v = 20 + 30 * (1 - (dist - max_) / (bad - max_))
    else:
        v = max(1, 20 * (1 - (dist - bad) / (bad + 1e-8)))
    return max(1, min(100, round(v)))


def _blend(dist_score, model_prob):
    """model_prob: 0-100 confidence the unimodal model gave the TRUE label
    for this attempt, or None if unavailable. Blends it with the
    distance-based score so a single attempt's own model confidence can
    correct for reference vectors being a poor/limited sample."""
    if model_prob is None:
        return dist_score
    blended = (1 - CONFIDENCE_BLEND_WEIGHT) * dist_score + CONFIDENCE_BLEND_WEIGHT * model_prob
    return max(1, min(100, round(blended)))


def score_attempt(audio_pooled, video_pooled, label, ref_stats_group, model_prob=None):
    """model_prob: the FUSION model's predicted probability (0-100) for
    `label` on THIS attempt — pass classify()'s true-label probability
    here. This is the only confidence signal available in production
    (only the fusion model ships to the app), so it's blended into both
    the audio and gesture distance-based scores. Optional; scoring falls
    back to pure distance if omitted."""
    if label not in ref_stats_group:
        return {"audio_score": None, "gesture_score": None, "combined_score": None}
    s = ref_stats_group[label]
    a = g = None
    if s.get("audio_vec") is not None:
        d = float(np.linalg.norm(audio_pooled - s["audio_vec"]))
        a = _blend(_score_one(d, s["audio_good"], s["audio_max"], s["audio_bad"]), model_prob)
    if s.get("lm_vec") is not None:
        d = float(np.linalg.norm(video_pooled - s["lm_vec"]))
        g = _blend(_score_one(d, s["lm_good"], s["lm_max"], s["lm_bad"]), model_prob)
    c = round(0.6 * a + 0.4 * g) if (a is not None and g is not None) else (a or g)
    return {"audio_score": a, "gesture_score": g, "combined_score": c}




# =============================================================================
# LABEL PARSING
# =============================================================================
def label_from_prefix(base):
    base = base.lower().replace(" ", "_")  # 'ai child.mp4' -> 'ai_child'
    for prefix in sorted(PREFIX_TO_LABEL, key=len, reverse=True):
        if base == prefix or base.startswith(prefix + "_") or base.startswith(prefix + "-"):
            return PREFIX_TO_LABEL[prefix]
    return None


def parse_test_filename(fname):
    """<label>_<number>_test_elder.mp4 or ..._test_child.mp4"""
    base = os.path.splitext(fname)[0].lower()
    group = None
    if base.endswith("_elder"):
        group = "elder"; base = base[:-6]
    elif base.endswith("_child"):
        group = "child"; base = base[:-6]
    if base.endswith("_test"):
        base = base[:-5]
    base = re.sub(r'_\d+$', '', base)
    return label_from_prefix(base), group


# =============================================================================
# BUILD (OR LOAD CACHED) REFERENCE STATS — elder + child
# =============================================================================
def compute_calibrated_thresholds(vecs_by_label, default_good, default_max, default_bad):
    """vecs_by_label: {label: pooled_vector}. For each label, finds the
    distance to its NEAREST different-class reference vector (the most
    confusable class) and derives good/max/bad from that gap:
        good = 35% of the nearest-confusable-class distance  (clearly correct)
        max_ = 75% of the nearest-confusable-class distance  (borderline)
        bad  = the farthest reference distance for that label (clearly wrong)
    This makes scoring self-scale to whatever the embedding space's actual
    distance range is, instead of assuming a fixed absolute distance —
    which is what was causing scores to be stuck low/near-zero after the
    real-sequence fix changed the pooled-vector distance scale.
    Falls back to the provided defaults when fewer than 2 classes exist
    (can't measure inter-class distance with only one reference).
    """
    labels = list(vecs_by_label.keys())
    thresholds = {}
    for label in labels:
        vec = vecs_by_label[label]
        dists = sorted(float(np.linalg.norm(vec - vecs_by_label[o]))
                        for o in labels if o != label)
        if not dists:
            thresholds[label] = (default_good, default_max, default_bad)
            continue
        nearest = dists[0]
        good = max(nearest * 0.35, 1e-6)
        max_ = max(nearest * 0.75, good * 1.5)
        bad = max(dists[-1], max_ * 1.5)
        thresholds[label] = (good, max_, bad)
    return thresholds


def build_ref_group(ref_dir):
    raw = {}
    if not os.path.isdir(ref_dir):
        print(f"  WARNING: {ref_dir} not found")
        return {}
    for fname in sorted(os.listdir(ref_dir)):
        if not fname.lower().endswith(".mp4"):
            continue
        label = label_from_prefix(os.path.splitext(fname)[0])
        if not label:
            print(f"  WARNING: cannot map '{fname}' to a class")
            continue
        path = os.path.join(ref_dir, fname)
        print(f"  Building ref stats: {fname} -> {label}")
        try:
            wav_tmp = os.path.join(TMP_DIR, "ref.wav")
            _mp4_to_clean_wav(path, wav_tmp)
            audio_seq = extract_audio_sequence(wav_tmp).numpy()
            video_seq = extract_video_sequence(path)
            raw[label] = {
                "audio_vec": audio_seq.mean(axis=0),   # pooled, for scoring
                "lm_vec": video_seq.mean(axis=0),       # pooled, for scoring
            }
        except Exception as e:
            print(f"  ERROR: {e}")

    if not raw:
        return {}

    audio_vecs = {l: v["audio_vec"] for l, v in raw.items() if v.get("audio_vec") is not None}
    lm_vecs = {l: v["lm_vec"] for l, v in raw.items() if v.get("lm_vec") is not None}
    audio_thr = compute_calibrated_thresholds(
        audio_vecs, DEFAULT_AUDIO_GOOD, DEFAULT_AUDIO_MAX, DEFAULT_AUDIO_BAD)
    lm_thr = compute_calibrated_thresholds(
        lm_vecs, DEFAULT_LM_GOOD, DEFAULT_LM_MAX, DEFAULT_LM_BAD)

    stats = {}
    for label, v in raw.items():
        ag, amax, abad = audio_thr.get(label, (DEFAULT_AUDIO_GOOD, DEFAULT_AUDIO_MAX, DEFAULT_AUDIO_BAD))
        lg, lmax, lbad = lm_thr.get(label, (DEFAULT_LM_GOOD, DEFAULT_LM_MAX, DEFAULT_LM_BAD))
        stats[label] = {
            "audio_vec": v.get("audio_vec"),
            "lm_vec": v.get("lm_vec"),
            "audio_good": ag, "audio_max": amax, "audio_bad": abad,
            "lm_good": lg, "lm_max": lmax, "lm_bad": lbad,
        }
    return stats


def _report_group_coverage(name, group_stats):
    missing = [l for l in LABELS_SORTED if l not in group_stats]
    print(f"  {name}: {len(group_stats)}/{len(LABELS_SORTED)} classes have reference stats")
    if missing:
        print(f"    WARNING: {name} is missing reference recordings for: {missing}")
        print(f"    -> scores for these classes will be None until reference "
              f"videos for them exist in the {name} reference folder.")


def get_reference_stats():
    if (not REBUILD_REF_STATS) and os.path.exists(REF_STATS_PKL):
        print(f"Loading cached reference stats from {REF_STATS_PKL}")
        with open(REF_STATS_PKL, "rb") as f:
            ref_stats = pickle.load(f)
        _report_group_coverage("elder", ref_stats.get("elder", {}))
        _report_group_coverage("child", ref_stats.get("child", {}))
        return ref_stats
    print("\nBuilding reference stats (elder)...")
    elder = build_ref_group(REF_ELDER_DIR)
    print("\nBuilding reference stats (child)...")
    child = build_ref_group(REF_CHILD_DIR)
    ref_stats = {"elder": elder, "child": child}
    print()
    _report_group_coverage("elder", elder)
    _report_group_coverage("child", child)
    with open(REF_STATS_PKL, "wb") as f:
        pickle.dump(ref_stats, f)
    print(f"✓ reference_stats.pkl saved -> {REF_STATS_PKL}")
    return ref_stats


# =============================================================================
# EXCEL BUILDER
# =============================================================================
def build_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation"

    BLUE, GREEN, RED, YELLOW = (PatternFill("solid", fgColor=c) for c in
                                 ["1B4F91", "D1FAE5", "FEE2E2", "FEF9C3"])
    ALT, WHITE = PatternFill("solid", fgColor="F8FAFC"), PatternFill("solid", fgColor="FFFFFF")

    def border():
        s = Side(style="thin", color="CBD5E1")
        return Border(left=s, right=s, top=s, bottom=s)

    headers = ["File", "Group", "True Label",
               "Fusion Predicted", "Fusion Confidence (%)", "Fusion True Label Prob (%)",
               "Audio-Only Predicted", "Audio-Only Confidence (%)", "Audio-Only True Label Prob (%)",
               "Video-Only Predicted", "Video-Only Confidence (%)", "Video-Only True Label Prob (%)",
               "Audio Score", "Gesture Score", "Overall Score",
               "Expert: Pronunciation (0-10)", "Expert: Gesture (0-10)",
               "Expert: Overall (0-10)", "Expert: Comments"]
    widths = [30, 8, 26,
              26, 18, 20,
              26, 20, 24,
              26, 20, 24,
              14, 14, 14, 24, 22, 22, 30]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = BLUE
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 36

    for r, row in enumerate(rows, start=2):
        ai_correct = row["true_label"] == row["pred_label"]
        row_fill = GREEN if ai_correct else RED
        alt_fill = ALT if r % 2 == 0 else WHITE

        def cell(col, val, fill=alt_fill, bold=False, color="111827"):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.font = Font(bold=bold, color=color, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border()

        ao_correct = row.get("audio_only_pred_label") is not None and \
                     row["true_label"] == row["audio_only_pred_label"]
        vo_correct = row.get("video_only_pred_label") is not None and \
                     row["true_label"] == row["video_only_pred_label"]

        cell(1, row["filename"], fill=row_fill)
        cell(2, row["group"].upper() if row["group"] else "?")
        cell(3, row["true_label"], fill=row_fill)

        cell(4, row["pred_label"], fill=row_fill, bold=True,
             color="1E3A5F" if ai_correct else "991B1B")
        cell(5, row["confidence"])
        cell(6, row["true_label_prob"])

        cell(7, row.get("audio_only_pred_label"), bold=True,
             color=("1E3A5F" if ao_correct else "991B1B") if row.get("audio_only_pred_label") else "111827")
        cell(8, row.get("audio_only_confidence"))
        cell(9, row.get("audio_only_true_label_prob"))

        cell(10, row.get("video_only_pred_label"), bold=True,
             color=("1E3A5F" if vo_correct else "991B1B") if row.get("video_only_pred_label") else "111827")
        cell(11, row.get("video_only_confidence"))
        cell(12, row.get("video_only_true_label_prob"))

        cell(13, row["audio_score"])
        cell(14, row["gesture_score"])
        cell(15, row["combined_score"], bold=True)
        for col in [16, 17, 18, 19]:
            c = ws.cell(row=r, column=col, value=None)
            c.fill = YELLOW
            c.border = border()

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    ws2 = wb.create_sheet("Summary")
    n = len(rows)
    n_correct = sum(1 for r in rows if r["true_label"] == r["pred_label"])
    ws2.append(["Metric", "Value"])
    ws2.append(["Total samples", n])
    ws2.append(["AI correct", n_correct])
    ws2.append(["AI accuracy (%)", round(n_correct / n * 100, 2) if n else 0])
    ws2.append([])
    ws2.append(["Per-group:"])
    for grp in ["elder", "child"]:
        g_rows = [r for r in rows if r["group"] == grp]
        g_corr = sum(1 for r in g_rows if r["true_label"] == r["pred_label"])
        pct = round(g_corr / len(g_rows) * 100, 1) if g_rows else 0
        ws2.append([f"  {grp} — accuracy", f"{g_corr}/{len(g_rows)} ({pct}%)"])
    for col in [1, 2]:
        ws2.column_dimensions[get_column_letter(col)].width = 40

    wb.save(out_path)
    print(f"\n✓ Excel saved -> {out_path}")
    print(f"  {n} samples | AI accuracy: {round(n_correct / n * 100, 2) if n else 0}%")


# =============================================================================
# STEP A: EXPERT EVALUATION
# =============================================================================
def run_expert_evaluation(ref_stats):
    test_files = sorted(f for f in os.listdir(TEST_DIR) if f.lower().endswith(".mp4"))
    print(f"\nFound {len(test_files)} test files in {TEST_DIR}")

    rows = []
    for fname in test_files:
        true_label, group = parse_test_filename(fname)
        if not true_label:
            print(f"  WARNING: cannot parse label from '{fname}' — skipping"); continue
        if not group:
            print(f"  WARNING: cannot parse group (elder/child) from '{fname}' — skipping"); continue

        path = os.path.join(TEST_DIR, fname)
        print(f"\n  {fname}  [{group}] -> true: {true_label}")

        try:
            wav_tmp = os.path.join(TMP_DIR, "test.wav")
            _mp4_to_clean_wav(path, wav_tmp)
            audio_seq = extract_audio_sequence(wav_tmp)          # (T, 768) real sequence
            video_seq = extract_video_sequence(path)              # (64, 234) real sequence
        except Exception as e:
            print(f"    ERROR extracting: {e}"); continue

        pred_label, confidence, class_probs = classify(audio_seq, video_seq)
        true_prob = round(class_probs.get(true_label, 0.0), 2)

        # unimodal predictions (blank if that checkpoint wasn't loaded)
        ao_pred, ao_conf, ao_probs = classify_audio_only(audio_seq)
        ao_true_prob = round(ao_probs.get(true_label, 0.0), 2) if ao_probs else None

        vo_pred, vo_conf, vo_probs = classify_video_only(video_seq)
        vo_true_prob = round(vo_probs.get(true_label, 0.0), 2) if vo_probs else None

        sc = score_attempt(audio_seq.numpy().mean(axis=0), video_seq.mean(axis=0),
                            true_label, ref_stats[group], model_prob=true_prob)

        print(f"    fusion pred={pred_label} ({confidence}%)  true_prob={true_prob}%  "
              f"audio={sc['audio_score']}  gesture={sc['gesture_score']}  combined={sc['combined_score']}")
        print(f"    audio-only pred={ao_pred} ({ao_conf}%)  true_prob={ao_true_prob}%   "
              f"video-only pred={vo_pred} ({vo_conf}%)  true_prob={vo_true_prob}%")

        rows.append({
            "filename": fname, "group": group, "true_label": true_label,
            "pred_label": pred_label, "confidence": confidence, "true_label_prob": true_prob,
            "audio_only_pred_label": ao_pred, "audio_only_confidence": ao_conf,
            "audio_only_true_label_prob": ao_true_prob,
            "video_only_pred_label": vo_pred, "video_only_confidence": vo_conf,
            "video_only_true_label_prob": vo_true_prob,
            "audio_score": sc["audio_score"], "gesture_score": sc["gesture_score"],
            "combined_score": sc["combined_score"],
        })

    if not rows:
        print("No results — check TEST_DIR and filename format")
        return
    rows.sort(key=lambda r: (r["group"], r["true_label"], r["filename"]))
    build_excel(rows, EXCEL_OUT)


# =============================================================================
# STEP B: EXPORT APP-INTEGRATION BUNDLE
# =============================================================================
INFERENCE_PY_TEMPLATE = r'''"""
inference.py — self-contained inference + scoring module for the
phoneme-gesture recognition model. Copy this file and the model_export/
folder into your app backend.

Usage:
    from inference import PhonemeGesturePredictor
    predictor = PhonemeGesturePredictor("/path/to/model_export")

    result = predictor.predict(wav_path, mp4_or_landmark_path)
    # {"predicted_class": str, "confidence": float, "class_probabilities": {...}}

    scores = predictor.score(wav_path, mp4_or_landmark_path, true_label, group)
    # {"audio_score": int|None, "gesture_score": int|None, "combined_score": int|None}
    # group must be "elder" or "child"
"""

import os, json, pickle
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import torchaudio
import cv2
import soundfile as sf
import noisereduce as nr
import mediapipe as mp
from mediapipe.tasks import python as mp_python
from mediapipe.tasks.python import vision as mp_vision
from transformers import WhisperProcessor, WhisperModel

WHISPER_LAYERS   = __WHISPER_LAYERS__
VIDEO_FRAMES     = __VIDEO_FRAMES__
MAX_AUDIO_FRAMES = __MAX_AUDIO_FRAMES__
VIDEO_FEAT_DIM   = __VIDEO_FEAT_DIM__
MOUTH_IDS = __MOUTH_IDS__


class AudioBiLSTMEncoder(nn.Module):
    def __init__(self, input_dim=768, hidden_size=512, num_layers=2, dropout=0.5):
        super().__init__()
        self.bilstm = nn.LSTM(input_dim, hidden_size, num_layers, bidirectional=True,
                               batch_first=True, dropout=dropout if num_layers > 1 else 0)
        self.ln = nn.LayerNorm(hidden_size * 2)
        self.attention = nn.Sequential(nn.Linear(hidden_size * 2, hidden_size), nn.Tanh(),
                                        nn.Linear(hidden_size, 1, bias=False))
        self.out_dim = hidden_size * 2

    def forward(self, x, lengths):
        packed = nn.utils.rnn.pack_padded_sequence(x, lengths.cpu(), batch_first=True, enforce_sorted=False)
        out, _ = self.bilstm(packed)
        out, _ = nn.utils.rnn.pad_packed_sequence(out, batch_first=True)
        seq = self.ln(out)
        w = F.softmax(self.attention(seq).squeeze(-1), dim=1)
        return seq, torch.bmm(w.unsqueeze(1), seq).squeeze(1)


class VideoBiLSTMEncoder(nn.Module):
    def __init__(self, input_dim=VIDEO_FEAT_DIM, hidden_size=128, num_layers=1, dropout=0.3):
        super().__init__()
        self.bilstm = nn.LSTM(input_dim, hidden_size, num_layers, bidirectional=True,
                               batch_first=True, dropout=dropout if num_layers > 1 else 0)
        self.ln = nn.LayerNorm(hidden_size * 2)
        self.attention = nn.Sequential(nn.Linear(hidden_size * 2, hidden_size), nn.Tanh(),
                                        nn.Linear(hidden_size, 1, bias=False))
        self.out_dim = hidden_size * 2

    def forward(self, x, lengths=None):
        out, _ = self.bilstm(x)
        seq = self.ln(out)
        w = F.softmax(self.attention(seq).squeeze(-1), dim=1)
        return seq, torch.bmm(w.unsqueeze(1), seq).squeeze(1)


class CrossAttentionFusion(nn.Module):
    def __init__(self, audio_dim, video_dim, proj_dim=256, num_heads=4, dropout=0.2):
        super().__init__()
        self.a_proj = nn.Linear(audio_dim, proj_dim)
        self.v_proj = nn.Linear(video_dim, proj_dim)
        self.a2v_attn = nn.MultiheadAttention(proj_dim, num_heads, dropout=dropout, batch_first=True)
        self.v2a_attn = nn.MultiheadAttention(proj_dim, num_heads, dropout=dropout, batch_first=True)
        self.ln_a = nn.LayerNorm(proj_dim)
        self.ln_v = nn.LayerNorm(proj_dim)
        self.out_dim = proj_dim * 2

    @staticmethod
    def _masked_mean(x, mask):
        valid = (~mask).unsqueeze(-1).float()
        return (x * valid).sum(1) / valid.sum(1).clamp(min=1)

    def forward(self, audio_seq, audio_mask, video_seq, video_mask):
        a = self.a_proj(audio_seq); v = self.v_proj(video_seq)
        v_ctx, _ = self.v2a_attn(query=v, key=a, value=a, key_padding_mask=audio_mask)
        a_ctx, _ = self.a2v_attn(query=a, key=v, value=v, key_padding_mask=video_mask)
        v_ctx = self.ln_v(v_ctx + v); a_ctx = self.ln_a(a_ctx + a)
        return torch.cat([self._masked_mean(a_ctx, audio_mask), self._masked_mean(v_ctx, video_mask)], dim=-1)


class FusionClassifier(nn.Module):
    def __init__(self, num_classes, proj_dim=256, dropout=0.3):
        super().__init__()
        self.audio_enc = AudioBiLSTMEncoder(hidden_size=512, num_layers=2, dropout=0.5)
        self.video_enc = VideoBiLSTMEncoder(hidden_size=128, num_layers=1)
        self.cross_attn = CrossAttentionFusion(self.audio_enc.out_dim, self.video_enc.out_dim,
                                                proj_dim=proj_dim, dropout=dropout)
        fused_dim = self.cross_attn.out_dim
        self.classifier = nn.Sequential(nn.Linear(fused_dim, 256), nn.ReLU(), nn.Dropout(dropout),
                                         nn.Linear(256, num_classes))

    def forward(self, video_x, video_len, audio_x, audio_len):
        v_seq, _ = self.video_enc(video_x)
        a_seq, _ = self.audio_enc(audio_x, audio_len)
        Tv, Ta = v_seq.shape[1], a_seq.shape[1]
        v_mask = torch.arange(Tv, device=video_x.device).unsqueeze(0) >= video_len.to(video_x.device).unsqueeze(1)
        a_mask = torch.arange(Ta, device=audio_x.device).unsqueeze(0) >= audio_len.to(audio_x.device).unsqueeze(1)
        return self.classifier(self.cross_attn(a_seq, a_mask, v_seq, v_mask))


class PhonemeGesturePredictor:
    def __init__(self, export_dir, device=None):
        self.device = device or ("cuda" if torch.cuda.is_available() else "cpu")
        self.export_dir = export_dir

        with open(os.path.join(export_dir, "label_map.json")) as f:
            maps = json.load(f)
        self.label_to_id = maps["label_to_id"]
        self.id_to_label = {int(k): v for k, v in maps["id_to_label"].items()}

        ref_path = os.path.join(export_dir, "reference_stats.pkl")
        self.reference_stats = {}
        if os.path.exists(ref_path):
            with open(ref_path, "rb") as f:
                self.reference_stats = pickle.load(f)  # {"elder": {...}, "child": {...}}

        self.processor = WhisperProcessor.from_pretrained(
            "openai/whisper-small", cache_dir=os.path.join(export_dir, "hf_cache"))
        whisper = WhisperModel.from_pretrained(
            "openai/whisper-small", output_hidden_states=True,
            cache_dir=os.path.join(export_dir, "hf_cache"))
        whisper.eval()
        self.whisper = whisper.to(self.device)
        self.layer_weights = torch.ones(len(WHISPER_LAYERS), device=self.device)
        self.layer_norm = nn.LayerNorm(768).to(self.device)

        self.model = FusionClassifier(num_classes=len(self.label_to_id)).to(self.device)
        ckpt = torch.load(os.path.join(export_dir, "model.pth"), map_location=self.device)
        self.model.load_state_dict(ckpt)
        self.model.eval()

        self._mp_dir = os.path.join(export_dir, "mp_models")
        os.makedirs(self._mp_dir, exist_ok=True)
        self._ensure_mp_models()
        print(f"Model loaded on {self.device} — {len(self.label_to_id)} classes")

    def _ensure_mp_models(self):
        import urllib.request
        files = {
            "hand_landmarker.task": "https://storage.googleapis.com/mediapipe-models/hand_landmarker/hand_landmarker/float16/1/hand_landmarker.task",
            "pose_landmarker_lite.task": "https://storage.googleapis.com/mediapipe-models/pose_landmarker/pose_landmarker_lite/float16/1/pose_landmarker_lite.task",
            "face_landmarker.task": "https://storage.googleapis.com/mediapipe-models/face_landmarker/face_landmarker/float16/1/face_landmarker.task",
        }
        for name, url in files.items():
            p = os.path.join(self._mp_dir, name)
            if not os.path.exists(p):
                urllib.request.urlretrieve(url, p)

    # ---- audio ----
    def _clean_wav(self, mp4_or_wav_path, tmp_wav):
        if mp4_or_wav_path.lower().endswith(".wav"):
            audio, sr = sf.read(mp4_or_wav_path)
        else:
            import subprocess
            subprocess.run(["ffmpeg", "-y", "-i", mp4_or_wav_path, "-vn", "-acodec", "pcm_s16le",
                             "-ar", "16000", "-ac", "1", tmp_wav],
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            audio, sr = sf.read(tmp_wav)
        audio = nr.reduce_noise(y=audio, sr=sr, stationary=True)
        sf.write(tmp_wav, audio, sr)
        return tmp_wav

    def _extract_audio_sequence(self, wav_path):
        wav, sr = torchaudio.load(wav_path)
        if wav.shape[0] > 1:
            wav = wav.mean(0, keepdim=True)
        if sr != 16000:
            wav = torchaudio.functional.resample(wav, sr, 16000)
        inputs = self.processor(wav.squeeze().numpy(), sampling_rate=16000, return_tensors="pt")
        feats = inputs.input_features.to(self.device)
        with torch.no_grad():
            out = self.whisper.encoder(input_features=feats)
            hidden = out.hidden_states
        selected = [hidden[i].squeeze(0) for i in WHISPER_LAYERS]
        w = F.softmax(self.layer_weights, dim=0)
        combined = torch.sum(w[:, None, None] * torch.stack(selected), dim=0)
        seq = self.layer_norm(combined).detach()
        if seq.shape[0] > MAX_AUDIO_FRAMES:
            seq = seq[:MAX_AUDIO_FRAMES]
        return seq.cpu()

    # ---- video ----
    def _extract_video_sequence(self, video_path):
        hl = mp_vision.HandLandmarker.create_from_options(
            mp_vision.HandLandmarkerOptions(
                base_options=mp_python.BaseOptions(model_asset_path=os.path.join(self._mp_dir, "hand_landmarker.task")),
                running_mode=mp_vision.RunningMode.VIDEO, num_hands=2))
        pl = mp_vision.PoseLandmarker.create_from_options(
            mp_vision.PoseLandmarkerOptions(
                base_options=mp_python.BaseOptions(model_asset_path=os.path.join(self._mp_dir, "pose_landmarker_lite.task")),
                running_mode=mp_vision.RunningMode.VIDEO))
        fl = mp_vision.FaceLandmarker.create_from_options(
            mp_vision.FaceLandmarkerOptions(
                base_options=mp_python.BaseOptions(model_asset_path=os.path.join(self._mp_dir, "face_landmarker.task")),
                running_mode=mp_vision.RunningMode.VIDEO))

        cap = cv2.VideoCapture(video_path)
        fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
        rows, idx = [], 0
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            mp_img = mp.Image(image_format=mp.ImageFormat.SRGB, data=rgb)
            ts = int(idx / fps * 1000)
            hr = hl.detect_for_video(mp_img, ts); pr = pl.detect_for_video(mp_img, ts); fr = fl.detect_for_video(mp_img, ts)
            hv = np.zeros((2, 21, 3), np.float32)
            for i, h_ in enumerate(hr.hand_landmarks[:2]):
                for j, lm in enumerate(h_):
                    hv[i, j] = [lm.x, lm.y, lm.z]
            pv = np.zeros((6, 3), np.float32)
            if pr.pose_landmarks:
                for i, pid in enumerate([11, 12, 13, 14, 15, 16]):
                    lm = pr.pose_landmarks[0][pid]; pv[i] = [lm.x, lm.y, lm.z]
            mv = np.zeros((len(MOUTH_IDS), 3), np.float32)
            if fr.face_landmarks:
                for i, fid in enumerate(MOUTH_IDS):
                    lm = fr.face_landmarks[0][fid]; mv[i] = [lm.x, lm.y, lm.z]
            rows.append(np.concatenate([hv.flatten(), pv.flatten(), mv.flatten()]))
            idx += 1
        cap.release()
        for x in [hl, pl, fl]:
            x.close()

        arr = np.stack(rows) if rows else np.zeros((1, VIDEO_FEAT_DIM), np.float32)
        T, D = arr.shape
        if T != VIDEO_FRAMES:
            src = np.linspace(0, T - 1, T); tgt = np.linspace(0, T - 1, VIDEO_FRAMES)
            out = np.zeros((VIDEO_FRAMES, D), np.float32)
            for d in range(D):
                out[:, d] = np.interp(tgt, src, arr[:, d])
            arr = out
        pose = arr[:, 126:144].reshape(VIDEO_FRAMES, 6, 3)
        mid = (pose[:, 0] + pose[:, 1]) / 2.0
        width = np.linalg.norm((pose[:, 0] - pose[:, 1])[:, :2], axis=1).clip(1e-4)
        out = arr.reshape(VIDEO_FRAMES, D // 3, 3)
        out = (out - mid[:, None, :]) / width[:, None, None]
        return out.reshape(VIDEO_FRAMES, D).astype(np.float32)

    def _load_landmarks_npy(self, lm_path):
        lm = np.load(lm_path).astype(np.float32)
        T, D = lm.shape
        pose = lm[:, 126:144].reshape(T, 6, 3)
        mid = (pose[:, 0] + pose[:, 1]) / 2.0
        width = np.linalg.norm((pose[:, 0] - pose[:, 1])[:, :2], axis=1).clip(1e-4)
        out = lm.reshape(T, D // 3, 3)
        out = (out - mid[:, None, :]) / width[:, None, None]
        lm = out.reshape(T, D).astype(np.float32)
        if T > VIDEO_FRAMES:
            lm = lm[:VIDEO_FRAMES]
        elif T < VIDEO_FRAMES:
            lm = np.vstack([lm, np.zeros((VIDEO_FRAMES - T, D), np.float32)])
        return lm

    def _get_sequences(self, audio_path, video_path_or_npy):
        tmp_wav = os.path.join("/tmp", "predict_tmp.wav")
        wav_path = self._clean_wav(audio_path, tmp_wav)
        audio_seq = self._extract_audio_sequence(wav_path)
        if video_path_or_npy.lower().endswith(".npy"):
            video_seq = self._load_landmarks_npy(video_path_or_npy)
        else:
            video_seq = self._extract_video_sequence(video_path_or_npy)
        return audio_seq, video_seq

    def predict(self, audio_path, video_path_or_npy):
        """audio_path: .wav or .mp4 (audio extracted). video_path_or_npy: .mp4 or pre-extracted (T,234) .npy"""
        audio_seq, video_seq = self._get_sequences(audio_path, video_path_or_npy)
        audio_x = audio_seq.unsqueeze(0).to(self.device)
        video_x = torch.from_numpy(video_seq).float().unsqueeze(0).to(self.device)
        audio_len = torch.tensor([audio_x.shape[1]])
        video_len = torch.tensor([video_x.shape[1]])
        with torch.no_grad():
            logits = self.model(video_x, video_len, audio_x, audio_len)
            probs = torch.sigmoid(logits).squeeze(0).cpu().numpy()
        pred_id = int(probs.argmax())
        class_probs = {self.id_to_label[i]: round(float(p) * 100, 2) for i, p in enumerate(probs)}
        return {
            "predicted_class": self.id_to_label[pred_id],
            "confidence": round(float(probs[pred_id]) * 100, 2),
            "class_probabilities": class_probs,
        }

    @staticmethod
    def _score_one(dist, good, max_, bad):
        if dist <= good:
            v = 80 + 20 * (1 - dist / good)
        elif dist <= max_:
            v = 50 + 30 * (1 - (dist - good) / (max_ - good))
        elif dist <= bad:
            v = 20 + 30 * (1 - (dist - max_) / (bad - max_))
        else:
            v = max(1, 20 * (1 - (dist - bad) / (bad + 1e-8)))
        return max(1, min(100, round(v)))

    CONFIDENCE_BLEND_WEIGHT = 0.4  # 0 = pure distance score, 1 = pure model confidence

    @classmethod
    def _blend(cls, dist_score, model_prob):
        """model_prob: the fusion model's own confidence (0-100) in the true
        label for THIS attempt, or None. Blended with the distance-based
        score since, in production, there's no distribution of other
        attempts to calibrate against — the model's own probability output
        on this single attempt is the only other independent signal
        available."""
        if model_prob is None:
            return dist_score
        blended = (1 - cls.CONFIDENCE_BLEND_WEIGHT) * dist_score + cls.CONFIDENCE_BLEND_WEIGHT * model_prob
        return max(1, min(100, round(blended)))

    def score(self, audio_path, video_path_or_npy, true_label, group="child"):
        """group: 'elder' or 'child'. Returns audio/gesture/combined 0-100
        scores against that group's reference recordings for true_label,
        blended with the fusion model's own confidence for this attempt."""
        group_stats = self.reference_stats.get(group, {})
        if true_label not in group_stats:
            return {"audio_score": None, "gesture_score": None, "combined_score": None}
        audio_seq, video_seq = self._get_sequences(audio_path, video_path_or_npy)

        # fusion model's confidence in the true label, for blending
        audio_x = audio_seq.unsqueeze(0).to(self.device)
        video_x = torch.from_numpy(video_seq).float().unsqueeze(0).to(self.device)
        audio_len = torch.tensor([audio_x.shape[1]])
        video_len = torch.tensor([video_x.shape[1]])
        with torch.no_grad():
            logits = self.model(video_x, video_len, audio_x, audio_len)
            probs = torch.sigmoid(logits).squeeze(0).cpu().numpy()
        model_prob = round(float(probs[self.label_to_id[true_label]]) * 100, 2)

        audio_pooled = audio_seq.numpy().mean(axis=0)
        video_pooled = video_seq.mean(axis=0)
        s = group_stats[true_label]
        a = g = None
        if s.get("audio_vec") is not None:
            d = float(np.linalg.norm(audio_pooled - s["audio_vec"]))
            a = self._blend(self._score_one(d, s["audio_good"], s["audio_max"], s["audio_bad"]), model_prob)
        if s.get("lm_vec") is not None:
            d = float(np.linalg.norm(video_pooled - s["lm_vec"]))
            g = self._blend(self._score_one(d, s["lm_good"], s["lm_max"], s["lm_bad"]), model_prob)
        c = round(0.6 * a + 0.4 * g) if (a is not None and g is not None) else (a or g)
        return {"audio_score": a, "gesture_score": g, "combined_score": c}
'''


def export_app_bundle():
    print("\n" + "=" * 60)
    print("EXPORTING APP-INTEGRATION BUNDLE")
    print("=" * 60)

    dst_ckpt = os.path.join(EXPORT_DIR, "model.pth")
    shutil.copy2(CHECKPOINT_PATH, dst_ckpt)
    print(f"✓ Checkpoint copied -> {dst_ckpt}")

    with open(os.path.join(EXPORT_DIR, "label_map.json"), "w") as f:
        json.dump({"label_to_id": label_to_id, "id_to_label": id_to_label}, f, indent=2)
    print("✓ Label map saved")

    with open(os.path.join(EXPORT_DIR, "model_config.json"), "w") as f:
        json.dump(MODEL_CONFIG, f, indent=2)
    print("✓ Model config saved")

    # reference_stats.pkl already written by get_reference_stats() into EXPORT_DIR
    if os.path.exists(REF_STATS_PKL):
        print(f"✓ reference_stats.pkl already in export dir")
    else:
        print("  WARNING: reference_stats.pkl missing — run get_reference_stats() first")

    code = (INFERENCE_PY_TEMPLATE
            .replace("__WHISPER_LAYERS__", repr(WHISPER_LAYERS))
            .replace("__VIDEO_FRAMES__", repr(VIDEO_FRAMES))
            .replace("__MAX_AUDIO_FRAMES__", repr(MAX_AUDIO_FRAMES))
            .replace("__VIDEO_FEAT_DIM__", repr(VIDEO_FEAT_DIM))
            .replace("__MOUTH_IDS__", repr(MOUTH_IDS)))
    with open(os.path.join(EXPORT_DIR, "inference.py"), "w") as f:
        f.write(code)
    print("✓ inference.py written (predict + score)")

    print(f"""
=============================================================
EXPORT COMPLETE — {EXPORT_DIR}/
=============================================================
  model.pth              — trained weights
  model_config.json      — architecture config
  label_map.json         — class id <-> label string
  reference_stats.pkl    — elder + child reference vectors for scoring
  inference.py            — copy this + this dir to your app backend

USAGE IN APP BACKEND:
  from inference import PhonemeGesturePredictor
  predictor = PhonemeGesturePredictor("/path/to/model_export")
  result = predictor.predict(wav_or_mp4_path, mp4_or_landmarks_npy_path)
  scores = predictor.score(wav_or_mp4_path, mp4_or_landmarks_npy_path,
                            true_label="s first letter of gp 01", group="child")
=============================================================
""")


# =============================================================================
# MAIN
# =============================================================================
if __name__ == "__main__":
    ref_stats = get_reference_stats()          # elder + child, cached
    run_expert_evaluation(ref_stats)            # -> expert_evaluation.xlsx
    export_app_bundle()                         # -> model_export/ (ships to app)